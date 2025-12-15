"""
RVU Comparison Script
Scans all Excel files in the current directory and compares procedures/RVUs
against rvu_settings.yaml rules, generating a report for each file.
"""

import json
import os
import sys
import yaml
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def load_rvu_settings(settings_file='rvu_settings.yaml'):
    """Load RVU settings from YAML file."""
    # Handle PyInstaller bundled files
    if getattr(sys, 'frozen', False):
        # Running as compiled executable
        base_path = Path(sys._MEIPASS)
    else:
        # Running as script
        base_path = Path(__file__).parent.absolute()
    
    settings_path = base_path / settings_file
    
    # Also check current working directory as fallback
    if not settings_path.exists():
        cwd_path = Path.cwd() / settings_file
        if cwd_path.exists():
            settings_path = cwd_path
    
    try:
        with open(settings_path, 'r', encoding='utf-8') as f:
            settings = yaml.safe_load(f)
        return settings
    except FileNotFoundError:
        print(f"ERROR: {settings_file} not found")
        if getattr(sys, 'frozen', False):
            print(f"  Running as frozen executable")
            print(f"  Checked bundled location: {base_path / settings_file}")
            print(f"  sys._MEIPASS: {sys._MEIPASS}")
            # List contents for debugging
            try:
                files_in_bundle = list(Path(sys._MEIPASS).iterdir())
                print(f"  Files in bundle: {[f.name for f in files_in_bundle]}")
            except:
                pass
        else:
            print(f"  Checked: {settings_path}")
        print(f"  Also checked working directory: {Path.cwd() / settings_file}")
        sys.exit(1)
    except yaml.YAMLError as e:
        print(f"ERROR: Invalid YAML in {settings_file}: {e}")
        sys.exit(1)


def find_column_index(headers, target_names):
    """
    Find column index by exact match of target names (case-insensitive).
    Returns (index, matched_name) or (None, None) if not found.
    Only matches exact column names, not partial matches.
    """
    headers_lower = [str(h).lower().strip() if h else "" for h in headers]
    target_names_lower = [name.lower().strip() for name in target_names]
    
    for idx, header in enumerate(headers_lower):
        if header in target_names_lower:
            return idx, headers[idx] if idx < len(headers) else None
    
    return None, None


def check_procedure_match(procedure_text, rvu_table, classification_rules, direct_lookups):
    """
    Check what study type and RVU would be assigned to this procedure.
    Mirrors the match_study_type function from RVUCounter.pyw to ensure consistency.
    """
    if not procedure_text:
        return "Unknown", 0.0
    
    if rvu_table is None:
        rvu_table = {}
    if classification_rules is None:
        classification_rules = {}
    if direct_lookups is None:
        direct_lookups = {}
    
    procedure_lower = procedure_text.lower().strip()
    procedure_stripped = procedure_text.strip()
    
    # Check classification rules
    classification_match_name = None
    classification_match_rvu = None
    
    # FIRST: Check user-defined classification rules (highest priority)
    for study_type, rules_list in classification_rules.items():
        if not isinstance(rules_list, list):
            continue
        
        for rule in rules_list:
            required_keywords = rule.get("required_keywords", [])
            excluded_keywords = rule.get("excluded_keywords", [])
            any_of_keywords = rule.get("any_of_keywords", [])
            
            # Special case for "CT Spine": exclude only if ALL excluded keywords are present
            if study_type == "CT Spine" and excluded_keywords:
                all_excluded = all(keyword.lower() in procedure_lower for keyword in excluded_keywords)
                if all_excluded:
                    continue
            # For other rules: exclude if any excluded keyword is present
            elif excluded_keywords:
                any_excluded = any(keyword.lower() in procedure_lower for keyword in excluded_keywords)
                if any_excluded:
                    continue
            
            # Check if all required keywords are present
            required_match = True
            if required_keywords:
                required_match = all(keyword.lower() in procedure_lower for keyword in required_keywords)
            
            # Check if at least one of any_of_keywords is present
            any_of_match = True
            if any_of_keywords:
                any_of_match = any(keyword.lower() in procedure_lower for keyword in any_of_keywords)
            
            # Match if all required keywords are present AND (any_of_keywords match OR no any_of_keywords specified)
            if required_match and any_of_match:
                rvu = rvu_table.get(study_type, 0.0)
                classification_match_name = study_type
                classification_match_rvu = rvu
                break
        
        if classification_match_name:
            break
    
    # If classification rule matched, return it immediately
    if classification_match_name:
        return classification_match_name, classification_match_rvu
    
    # THIRD: Try exact match
    for study_type, rvu in rvu_table.items():
        if study_type.lower() == procedure_lower:
            return study_type, rvu
    
    # FOURTH: Try keyword matching (look up RVU values from rvu_table)
    keyword_study_types = {
        "ct cap": "CT CAP",
        "ct ap": "CT AP",
        "cta": "CTA Brain",  # Default CTA
        "ultrasound": "US Other",  # Check "ultrasound" before "us"
        "mri": "MRI Other",
        "mr ": "MRI Other",
        "us ": "US Other",
        "x-ray": "XR Other",
        "xr ": "XR Other",
        "xr\t": "XR Other",  # XR with tab
        "nuclear": "NM Other",
        "nm ": "NM Other",
    }
    
    for keyword in sorted(keyword_study_types.keys(), key=len, reverse=True):
        if keyword in procedure_lower:
            study_type = keyword_study_types[keyword]
            rvu = rvu_table.get(study_type, 0.0)
            return study_type, rvu
    
    # FIFTH: Check prefix (look up RVU values from rvu_table)
    # IMPORTANT: Check XA before CT (since "xa" starts with "x" which could match "xr")
    if len(procedure_lower) >= 2:
        first_two = procedure_lower[:2]
        # Check for 3-character prefixes first (XA, CTA) before 2-character
        if len(procedure_lower) >= 3:
            first_three = procedure_lower[:3]
            if first_three == "xa " or first_three == "xa\t":
                # XA is fluoroscopy (XR modality)
                return "XR Other", rvu_table.get("XR Other", 0.3)
            elif first_three == "cta":
                # CTA - will be handled by classification rules or keyword matching
                pass
        
        prefix_study_types = {
            "xr": "XR Other",
            "x-": "XR Other",
            "ct": "CT Other",
            "mr": "MRI Other",
            "us": "US Other",
            "nm": "NM Other",
        }
        if first_two in prefix_study_types:
            study_type = prefix_study_types[first_two]
            rvu = rvu_table.get(study_type, 0.0)
            return study_type, rvu
    
    # SIXTH: Try partial matches (most specific first)
    matches = []
    other_matches = []
    pet_ct_match = None
    
    for study_type, rvu in rvu_table.items():
        study_lower = study_type.lower()
        
        # Special handling for PET CT
        if study_lower == "pet ct":
            if "pet" in procedure_lower and "ct" in procedure_lower:
                pet_ct_match = (study_type, rvu)
            continue
        
        if study_lower in procedure_lower or procedure_lower in study_lower:
            score = len(study_type)
            if " other" in study_lower or study_lower.endswith(" other"):
                other_matches.append((score, study_type, rvu))
            else:
                matches.append((score, study_type, rvu))
    
    # Return most specific non-"Other" match if found
    if matches:
        matches.sort(reverse=True)  # Highest score first
        return matches[0][1], matches[0][2]
    
    # If no specific match, try "Other" types as fallback
    if other_matches:
        other_matches.sort(reverse=True)  # Highest score first
        return other_matches[0][1], other_matches[0][2]
    
    # Absolute last resort: PET CT
    if pet_ct_match:
        return pet_ct_match
    
    return "Unknown", 0.0


def process_excel_file(excel_path, settings):
    """Process a single Excel file and return comparison results."""
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        return {
            'error': f"Error opening file: {e}",
            'sheets': [],
            'total_procedures': 0,
            'outliers': []
        }
    
    rvu_table = settings.get('rvu_table', {})
    direct_lookups = settings.get('direct_lookups', {})
    classification_rules = settings.get('classification_rules', {})
    
    all_outliers = []
    sheet_results = []
    total_procedures = 0
    
    # First pass: Find which sheet(s) have the required columns
    sheets_with_columns = []
    all_sheets_checked = []
    
    for sheet_name in wb.sheetnames:
        # Skip "Summary" sheet (case-insensitive)
        if sheet_name.lower() == 'summary':
            continue
        ws = wb[sheet_name]
        
        # Find header row (first non-empty row)
        header_row = None
        headers = None
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if any(cell for cell in row):
                header_row = row_idx
                headers = list(row)
                break
        
        if header_row is None:
            sheet_results.append({
                'name': sheet_name,
                'error': 'No header row found',
                'procedures': 0,
                'outliers': []
            })
            all_sheets_checked.append(sheet_name)
            continue
        
        # Check for exact column matches (only StandardProcedureName and wRVU_Matrix)
        proc_col_idx, proc_col_name = find_column_index(
            headers, 
            ['StandardProcedureName']
        )
        rvu_col_idx, rvu_col_name = find_column_index(
            headers,
            ['wRVU_Matrix']
        )
        
        if proc_col_idx is not None and rvu_col_idx is not None:
            sheets_with_columns.append({
                'name': sheet_name,
                'worksheet': ws,
                'header_row': header_row,
                'headers': headers,
                'proc_col_idx': proc_col_idx,
                'proc_col_name': proc_col_name,
                'rvu_col_idx': rvu_col_idx,
                'rvu_col_name': rvu_col_name
            })
        else:
            # Sheet doesn't have required columns - report it
            missing = []
            if proc_col_idx is None:
                missing.append('StandardProcedureName')
            if rvu_col_idx is None:
                missing.append('wRVU_Matrix')
            sheet_results.append({
                'name': sheet_name,
                'error': f'Missing required columns: {", ".join(missing)}',
                'procedures': 0,
                'outliers': []
            })
        
        all_sheets_checked.append(sheet_name)
    
    # Process only sheets that have both required columns
    for sheet_info in sheets_with_columns:
        sheet_name = sheet_info['name']
        ws = sheet_info['worksheet']
        header_row = sheet_info['header_row']
        proc_col_idx = sheet_info['proc_col_idx']
        rvu_col_idx = sheet_info['rvu_col_idx']
        proc_col_name = sheet_info['proc_col_name']
        rvu_col_name = sheet_info['rvu_col_name']
        
        # Process data rows
        procedures_with_rvu = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if proc_col_idx < len(row) and rvu_col_idx < len(row):
                proc = row[proc_col_idx]
                rvu_val = row[rvu_col_idx]
                
                if proc and str(proc).strip():
                    proc_str = str(proc).strip()
                    rvu = None
                    if rvu_val is not None:
                        try:
                            rvu = float(rvu_val)
                        except (ValueError, TypeError):
                            pass
                    
                    if rvu is not None:
                        procedures_with_rvu.append((proc_str, rvu))
        
        # Compare procedures
        outliers = []
        for proc, excel_rvu in procedures_with_rvu:
            matched_type, matched_rvu = check_procedure_match(
                proc, rvu_table, classification_rules, direct_lookups
            )
            
            if matched_type is None:
                outliers.append({
                    'procedure': proc,
                    'excel_rvu': excel_rvu,
                    'matched_type': 'NO MATCH',
                    'matched_rvu': None,
                    'reason': 'No matching rule found'
                })
            elif abs(matched_rvu - excel_rvu) > 0.01:  # Allow small floating point differences
                outliers.append({
                    'procedure': proc,
                    'excel_rvu': excel_rvu,
                    'matched_type': matched_type,
                    'matched_rvu': matched_rvu,
                    'reason': f'RVU mismatch: Excel has {excel_rvu}, rules match to {matched_rvu}'
                })
        
        sheet_results.append({
            'name': sheet_name,
            'proc_column': proc_col_name,
            'rvu_column': rvu_col_name,
            'procedures': len(procedures_with_rvu),
            'outliers': outliers
        })
        
        all_outliers.extend(outliers)
        total_procedures += len(procedures_with_rvu)
    
    wb.close()
    
    return {
        'sheets': sheet_results,
        'total_procedures': total_procedures,
        'outliers': all_outliers
    }


def generate_report(excel_path, results, output_path):
    """Generate a text report for an Excel file."""
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("=" * 100 + "\n")
        f.write(f"RVU COMPARISON REPORT\n")
        f.write("=" * 100 + "\n")
        f.write(f"Excel File: {excel_path}\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 100 + "\n\n")
        
        if 'error' in results:
            f.write(f"ERROR: {results['error']}\n")
            return
        
        f.write(f"Total Procedures Processed: {results['total_procedures']}\n")
        f.write(f"Total Outliers Found: {len(results['outliers'])}\n\n")
        
        # Report by sheet
        for sheet_info in results['sheets']:
            f.write("-" * 100 + "\n")
            f.write(f"Sheet: {sheet_info['name']}\n")
            f.write("-" * 100 + "\n")
            
            if 'error' in sheet_info:
                f.write(f"  ERROR: {sheet_info['error']}\n\n")
                continue
            
            f.write(f"  Procedure Column: {sheet_info.get('proc_column', 'N/A')}\n")
            f.write(f"  RVU Column: {sheet_info.get('rvu_column', 'N/A')}\n")
            f.write(f"  Procedures in Sheet: {sheet_info['procedures']}\n")
            f.write(f"  Outliers in Sheet: {len(sheet_info['outliers'])}\n\n")
            
            if sheet_info['outliers']:
                # Group unique outliers
                unique_outliers = {}
                for outlier in sheet_info['outliers']:
                    key = outlier['procedure']
                    if key not in unique_outliers:
                        unique_outliers[key] = outlier
                
                f.write(f"  Unique Outliers ({len(unique_outliers)}):\n")
                for proc, outlier in sorted(unique_outliers.items()):
                    f.write(f"    Procedure: {outlier['procedure']}\n")
                    f.write(f"      Excel RVU: {outlier['excel_rvu']}\n")
                    f.write(f"      Matched Type: {outlier['matched_type']}\n")
                    if outlier['matched_rvu'] is not None:
                        f.write(f"      Matched RVU: {outlier['matched_rvu']}\n")
                    f.write(f"      Reason: {outlier['reason']}\n")
                    f.write("\n")
            else:
                f.write("  All procedures match the rules!\n\n")
        
        # Summary of all outliers
        if results['outliers']:
            f.write("\n" + "=" * 100 + "\n")
            f.write("SUMMARY OF ALL OUTLIERS\n")
            f.write("=" * 100 + "\n\n")
            
            # Group by unique procedure
            unique_outliers = {}
            for outlier in results['outliers']:
                key = outlier['procedure']
                if key not in unique_outliers:
                    unique_outliers[key] = outlier
            
            f.write(f"Unique Outlier Procedures: {len(unique_outliers)}\n\n")
            for proc, outlier in sorted(unique_outliers.items()):
                f.write(f"  {outlier['procedure']}\n")
                f.write(f"    Excel RVU: {outlier['excel_rvu']}\n")
                f.write(f"    Matched Type: {outlier['matched_type']}\n")
                if outlier['matched_rvu'] is not None:
                    f.write(f"    Matched RVU: {outlier['matched_rvu']}\n")
                f.write(f"    Reason: {outlier['reason']}\n\n")
        else:
            f.write("\n" + "=" * 100 + "\n")
            f.write("SUCCESS: All procedures match the rules!\n")
            f.write("=" * 100 + "\n")


def main():
    """Main function to process all Excel files in current directory."""
    # When frozen, use the directory where the exe is run from
    # When running as script, use the script's directory
    if getattr(sys, 'frozen', False):
        # Running as compiled executable - use current working directory
        work_dir = Path.cwd()
    else:
        # Running as script - use script's directory
        work_dir = Path(__file__).parent.absolute()
        os.chdir(work_dir)
    
    print("RVU Excel File Comparison Tool")
    print("=" * 60)
    print(f"Working directory: {work_dir}\n")
    
    # Load settings
    print("Loading rvu_settings.yaml...")
    settings = load_rvu_settings()
    print("Settings loaded successfully.\n")
    
    # Find all Excel files in the working directory
    excel_files = list(work_dir.glob('*.xlsx')) + list(work_dir.glob('*.xlsm'))
    
    if not excel_files:
        print("No Excel files (.xlsx or .xlsm) found in current directory.")
        return
    
    print(f"Found {len(excel_files)} Excel file(s):")
    for f in excel_files:
        print(f"  - {f.name}")
    print()
    
    # Process each file
    for excel_file in excel_files:
        print(f"Processing: {excel_file.name}...")
        
        results = process_excel_file(excel_file, settings)
        
        # Generate report in the same directory as the Excel file
        report_file = work_dir / (excel_file.stem + '_rvu_report.txt')
        generate_report(excel_file.name, results, report_file)
        
        print(f"  Procedures: {results['total_procedures']}")
        print(f"  Outliers: {len(results['outliers'])}")
        print(f"  Report saved to: {report_file}\n")
    
    print("Processing complete!")


if __name__ == '__main__':
    main()

